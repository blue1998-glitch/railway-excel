import io
import os
import time
import openpyxl
import pandas as pd
from pydantic import BaseModel, Field
from pypdf import PdfReader, PdfWriter
import streamlit as st
from google import genai
from google.genai import types

# ----------------------------------------------------
# 1. 網頁基本設定與金鑰讀取
# ----------------------------------------------------
st.set_page_config(page_title="台鐵解款單自動化填表系統", page_icon="🚆", layout="wide")
st.title("🚆 台鐵掃描解款單 ➜ Excel 智慧自動填表系統")
st.caption("🚀 完整保留原始公式與格式 ｜ ⚖️ 會計立場扣抵公式 (=刷卡-退刷) ｜ 🔍 自動檢核自輸總計")

raw_key = st.secrets.get("GEMINI_API_KEY", "")
cleaned_key = str(raw_key).replace('"', '').replace("'", "").strip()

with st.sidebar:
    st.header("⚙️ 系統設定")
    user_key = st.text_input(
        "Gemini API Key",
        value=cleaned_key,
        type="password",
        help="系統會優先讀取 secrets 中的 GEMINI_API_KEY"
    )
    active_api_key = user_key.strip()
    
    # 動態探測此 API Key 支援的所有模型清單（保證不發生 404）
    available_models = []
    if active_api_key:
        try:
            temp_client = genai.Client(api_key=active_api_key)
            for m in temp_client.models.list():
                m_name = getattr(m, "name", "").replace("models/", "")
                if "gemini" in m_name.lower():
                    available_models.append(m_name)
        except Exception:
            pass

    if not available_models:
        available_models = [
            "gemini-2.0-flash",
            "gemini-1.5-flash",
            "gemini-1.5-pro"
        ]

    # 預設優先選擇 flash 系列（速度最快）
    default_idx = 0
    for i, name in enumerate(available_models):
        if "flash" in name.lower():
            default_idx = i
            break

    selected_model = st.selectbox(
        "AI 辨識核心模型",
        options=available_models,
        index=default_idx,
        help="系統已自動連線並篩選出您金鑰支援的模型"
    )

    if active_api_key:
        st.success("✅ API 金鑰連線正常")
    else:
        st.warning("⚠️ 請確認已在 secrets 設定或在此輸入 API Key")

    st.markdown("---")
    st.markdown("""
    💡 **會計計算原則**：
    * **電腦信用卡**：`信用卡刷卡(-)` 減 `信用卡退刷(+)`
    * **條碼**：`條碼支付進款(-)` 減 `條碼支付退款(+)`
    * **平衡檢核**：客運 + 貨運 - 電腦信用卡 - 條碼 + 其他 ＝ 自輸(應解總計)
    """)

# ----------------------------------------------------
# 2. 定義資料結構與工具函式
# ----------------------------------------------------
class StationReport(BaseModel):
    station_name: str = Field(description="車站名稱（例如：豐富、苗栗、銅鑼、三義、泰安、后里、豐原、栗林、潭子、頭家厝、松竹、太原、精武、台中、五權、大慶、新烏日、烏日、成功、彰化、花壇、大村、員林、社頭、田中、二水等，不需包含'站'字）")
    date_day: int = Field(description="報表日期中的『日/號』(1 至 31 的整數數字)")
    passenger_revenue: float = Field(default=0.0, description="左側【應解款數】區塊內的『客運(+)』金額，無則為 0")
    freight_revenue: float = Field(default=0.0, description="左側【應解款數】區塊內的『貨運(+)』金額，無則為 0")
    credit_card_charge: float = Field(default=0.0, description="左側【應解款數】區塊內的『信用卡刷卡(-)』金額 (填正數)，無則為 0")
    credit_card_refund: float = Field(default=0.0, description="左側【應解款數】區塊內的『信用卡退刷(+)』金額 (填正數)，無則為 0")
    barcode_in: float = Field(default=0.0, description="左側【應解款數】區塊內的『條碼支付進款(-)』金額 (填正數)，無則為 0")
    barcode_refund: float = Field(default=0.0, description="左側【應解款數】區塊內的『條碼支付退款(+)』金額 (填正數)，無則為 0")
    other_amount: float = Field(default=0.0, description="左側【應解款數】區塊內除上述項目外的其他明細金額加總（如存付運費、託收支票、補繳金額、繳回週轉金、其他短欠等），若無則填 0")
    remittance_total: float = Field(default=0.0, description="左側【應解款數】區塊內的『應解總計』金額")

def extract_json_str(text: str) -> str:
    """安全去除 Markdown 標籤以解析 JSON"""
    if not text:
        return "{}"
    t = text.strip()
    if t.startswith("```"):
        lines = t.splitlines()
        if lines and lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].startswith("```"):
            lines = lines[:-1]
        t = "\n".join(lines).strip()
    return t

def clean_station_name(val):
    """標準化車站名稱以利比對工作表名稱"""
    if not val:
        return ""
    return str(val).replace("臺", "台").replace("站", "").replace(" ", "").replace("　", "").strip()

def to_clean_num(val):
    """轉換為乾淨整數或保留小數"""
    try:
        f_val = float(val)
        return int(f_val) if f_val.is_integer() else f_val
    except Exception:
        return 0

def build_deduction_formula(charge_val, refund_val):
    """會計立場相減公式：=刷卡-退刷"""
    c = to_clean_num(charge_val)
    r = to_clean_num(refund_val)
    if c == 0 and r == 0:
        return None
    if c > 0 and r > 0:
        return f"={c}-{r}"
    if c > 0 and r == 0:
        return c
    if c == 0 and r > 0:
        return f"=-{r}"
    return None

def load_excel_preserving_all(file_bytes, filename):
    """載入 Excel 並全力保護原始公式與格式"""
    if filename.lower().endswith(".xls"):
        try:
            import xlrd
        except ImportError:
            raise ImportError("請在 requirements.txt 中加入 xlrd。")

        xls_book = xlrd.open_workbook(file_contents=file_bytes, formatting_info=False)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        for sheet_name in xls_book.sheet_names():
            xls_sheet = xls_book.sheet_by_name(sheet_name)
            xlsx_sheet = wb.create_sheet(title=sheet_name)
            for r in range(xls_sheet.nrows):
                for c in range(xls_sheet.ncols):
                    cell = xls_sheet.cell(r, c)
                    val = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            val = xlrd.xldate.xldate_as_datetime(val, xls_book.datemode)
                        except Exception:
                            pass
                    if val != "" and val is not None:
                        xlsx_sheet.cell(row=r + 1, column=c + 1, value=val)
            
            header_row = [str(xlsx_sheet.cell(1, c).value or "") for c in range(1, 10)]
            if any("客運" in h for h in header_row):
                for day_r in range(2, 33):
                    xlsx_sheet.cell(row=day_r, column=8, value=f"=B{day_r}+C{day_r}-D{day_r}-E{day_r}+F{day_r}")
                    xlsx_sheet.cell(row=day_r, column=9, value=f"=G{day_r}-H{day_r}")
        return wb
    else:
        return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)

def analyze_sheet_structure(sheet):
    """智慧掃描工作表表頭取得各欄位位置"""
    col_map = {}
    for r in range(1, 8):
        for c in range(1, sheet.max_column + 1):
            val = str(sheet.cell(row=r, column=c).value or "").replace(" ", "").replace("\n", "").strip()
            if not val:
                continue
            if "客運" in val and "passenger" not in col_map:
                col_map["passenger"] = c
            elif "貨運" in val and "freight" not in col_map:
                col_map["freight"] = c
            elif ("信用卡" in val or "刷卡" in val) and "credit" not in col_map:
                col_map["credit"] = c
            elif ("條碼" in val or "支付" in val) and "barcode" not in col_map:
                col_map["barcode"] = c
            elif "其他" in val and "other" not in col_map:
                col_map["other"] = c
            elif ("自輸" in val or "字輸" in val) and "remittance" not in col_map:
                col_map["remittance"] = c
            elif ("日" in val or "期" in val or "號" in val) and "date" not in col_map:
                col_map["date"] = c

    defaults = {
        "date": 1,
        "passenger": 2,
        "freight": 3,
        "credit": 4,
        "barcode": 5,
        "other": 6,
        "remittance": 7
    }
    for k, v in defaults.items():
        if k not in col_map:
            col_map[k] = v
    return col_map

def find_target_row(sheet, date_day, date_col=1):
    """精準尋找對應日期的列號 (1~31)"""
    for r in range(1, 50):
        val = sheet.cell(row=r, column=date_col).value
        if val is not None:
            try:
                val_str = str(val).replace("日", "").replace("號", "").strip()
                if int(float(val_str)) == int(date_day):
                    return r
            except Exception:
                pass
    return int(date_day) + 1

def write_cell_if_valid(sheet, row_idx, col_idx, val):
    """安全寫入指定儲存格（0 或空值不寫入）"""
    if val is not None and val != 0 and val != "0" and val != "":
        if col_idx:
            sheet.cell(row=row_idx, column=col_idx, value=val)
            return 1
    return 0

# ----------------------------------------------------
# 3. 介面上傳區塊
# ----------------------------------------------------
col1, col2 = st.columns(2)
with col1:
    uploaded_excel = st.file_uploader("📥 步驟 1：上傳 Excel 公版 (.xlsx 或 .xls)", type=["xlsx", "xls"])
with col2:
    uploaded_pdfs = st.file_uploader("📥 步驟 2：批次上傳掃描 PDF 解款單 (可多選)", type=["pdf"], accept_multiple_files=True)

# ----------------------------------------------------
# 4. 核心辨識與填寫
# ----------------------------------------------------
if st.button("🚀 開始智慧辨識與自動填表", type="primary", use_container_width=True):
    if not active_api_key:
        st.error("❌ 尚未設定 Gemini API Key，請在側邊欄輸入！")
        st.stop()
    if not uploaded_excel or not uploaded_pdfs:
        st.error("❌ 請確認已同時上傳 Excel 公版與 PDF 檔案！")
        st.stop()

    start_time = time.time()
    client = genai.Client(api_key=active_api_key)

    try:
        wb = load_excel_preserving_all(uploaded_excel.getvalue(), uploaded_excel.name)
    except Exception as e:
        st.error(f"❌ Excel 讀取失敗：{e}")
        st.stop()

    status_box = st.status("📄 [階段 1/2] 正在讀取與拆分 PDF 頁面...", expanded=True)
    all_pages = []
    
    for pdf_file in uploaded_pdfs:
        status_box.write(f"🔍 讀取檔案 `{pdf_file.name}` ...")
        try:
            reader = PdfReader(io.BytesIO(pdf_file.getvalue()))
            total_p = len(reader.pages)
            for p_idx, page in enumerate(reader.pages, 1):
                writer = PdfWriter()
                writer.add_page(page)
                page_buf = io.BytesIO()
                writer.write(page_buf)
                all_pages.append((pdf_file.name, p_idx, total_p, page_buf.getvalue()))
        except Exception as e:
            status_box.write(f"⚠️ 檔案 `{pdf_file.name}` 解析異常：{e}")

    total_tasks = len(all_pages)
    if total_tasks == 0:
        status_box.update(label="❌ 沒有找到可處理的 PDF 頁面", state="error")
        st.stop()

    status_box.update(label=f"🚀 [階段 2/2] 正在辨識 {total_tasks} 頁解款單據並填寫 Excel...", state="running")
    
    # 強化淺色印件識別與會計勾稽驗算
    prompt = """
    你是一位專業精確的台鐵會計表單辨識專家。請仔細檢視這張解款單據：
    【特別注意淺色印件】：
    本單據若為影印本，字跡、複寫痕跡或印泥顏色可能較淺，請特別仔細分辨淺色或模糊數字（如 0、3、8、1、7），切勿遺漏淺色墨跡。
    1. 擷取【車站名稱】（例如：豐富、苗栗、銅鑼、三義、泰安、后里、豐原、栗林、潭子、頭家厝、松竹、太原、精武、台中、五權、大慶、新烏日、烏日、成功、彰化、花壇、大村、員林、社頭、田中、二水等，不需包含'站'字）與進款【日期】（僅需日/號數，1-31 的整數）。
    2. 專注看左側【應解款數】大項目區塊，精確擷取各數值（若無填 0）：
       - 客運(+)
       - 貨運(+)
       - 信用卡刷卡(-) (填正數)
       - 信用卡退刷(+) (填正數)
       - 條碼支付進款(-) (填正數)
       - 條碼支付退款(+) (填正數)
       - 應解總計 (報表上的應解總計數值)
       - 其他項目加總（若有存付運費、託收支票、補繳金額、繳回週轉金、自售機短欠等款項，請計算其淨額填入其他；若無則填 0）
    3. 會計勾稽驗算原則：正常情況下必符合「客運 + 貨運 - 信用卡刷卡 + 信用卡退刷 - 條碼進款 + 條碼退款 + 其他 = 應解總計」。若遇到淺色不確定字跡，請務必利用此勾稽平衡原則反推驗算確認！
    """

    models_to_try = [selected_model] + [m for m in available_models if m != selected_model]

    results_data = []
    audit_records = []
    progress_bar = st.progress(0)
    
    for idx, (file_name, p_idx, total_p, page_bytes) in enumerate(all_pages, 1):
        progress_bar.progress(idx / total_tasks, text=f"正在辨識第 {idx}/{total_tasks} 頁...")
        
        parsed_data = None
        err_msg = ""
        
        # 輪詢探測可用模型；一旦某模型成功，立即將該模型調至第一位，後續頁面 100% 直通無延遲
        for model_name in list(models_to_try):
            success = False
            for retry in range(2):
                try:
                    res = client.models.generate_content(
                        model=model_name,
                        contents=[
                            types.Part.from_bytes(data=page_bytes, mime_type="application/pdf"),
                            prompt
                        ],
                        config=types.GenerateContentConfig(
                            response_mime_type="application/json",
                            response_schema=StationReport,
                            temperature=0.0,
                        )
                    )
                    if res and res.text:
                        clean_text = extract_json_str(res.text)
                        parsed_data = StationReport.model_validate_json(clean_text)
                        # 核心優化：記住成功模型，後續頁面直接置頂呼叫，不再試無效模型
                        if models_to_try[0] != model_name:
                            models_to_try.remove(model_name)
                            models_to_try.insert(0, model_name)
                        success = True
                        break
                except Exception as e:
                    err_msg = str(e)
                    if "429" in err_msg or "RESOURCE_EXHAUSTED" in err_msg:
                        time.sleep(4.0)
                        continue
                    else:
                        # 遇 404 或其他錯誤，直接換下一個模型測試
                        break
            if success:
                break

        if parsed_data and parsed_data.date_day > 0:
            results_data.append((file_name, p_idx, parsed_data))
            status_box.write(f"✅ [{idx:02d}/{total_tasks:02d}] **{parsed_data.station_name}**（{parsed_data.date_day} 日）辨識成功")
        else:
            status_box.write(f"⚠️ [{idx:02d}/{total_tasks:02d}] `{file_name}` 第 {p_idx} 頁辨識失敗 ({err_msg[:60]}...)")
        
        time.sleep(1.0)

    # ----------------------------------------------------
    # 5. 回填 Excel 與 會計立場平衡檢查
    # ----------------------------------------------------
    total_written = 0
    success_count = 0

    for file_name, p_idx, data in results_data:
        target_name_clean = clean_station_name(data.station_name)
        
        # 優先完全符合（防止「烏日」被「新烏日」搶先匹配）
        target_sheet = None
        for s_name in wb.sheetnames:
            if clean_station_name(s_name) == target_name_clean:
                target_sheet = wb[s_name]
                break
        if not target_sheet:
            for s_name in wb.sheetnames:
                s_clean = clean_station_name(s_name)
                if s_clean and (s_clean in target_name_clean or target_name_clean in s_clean):
                    target_sheet = wb[s_name]
                    break

        net_credit = data.credit_card_charge - data.credit_card_refund
        net_barcode = data.barcode_in - data.barcode_refund

        computed_total = data.passenger_revenue + data.freight_revenue - net_credit - net_barcode + data.other_amount
        diff = round(data.remittance_total - computed_total, 2)
        is_balanced = (abs(diff) < 0.01)

        audit_records.append({
            "檔案名稱": file_name,
            "車站名稱": data.station_name,
            "日期": f"{data.date_day} 日",
            "客運": to_clean_num(data.passenger_revenue),
            "貨運": to_clean_num(data.freight_revenue),
            "電腦信用卡 (=刷卡-退刷)": f"{to_clean_num(data.credit_card_charge)} - {to_clean_num(data.credit_card_refund)}" if data.credit_card_refund > 0 else to_clean_num(data.credit_card_charge),
            "條碼 (=進款-退款)": f"{to_clean_num(data.barcode_in)} - {to_clean_num(data.barcode_refund)}" if data.barcode_refund > 0 else to_clean_num(data.barcode_in),
            "其他": to_clean_num(data.other_amount),
            "自輸 (PDF應解總計)": to_clean_num(data.remittance_total),
            "計算總計": to_clean_num(computed_total),
            "差額 (自輸-總計)": diff,
            "平衡狀態": "✅ 平衡 (0)" if is_balanced else f"❌ 差額 {diff:+.0f} (請核對)",
            "工作表寫入": f"已寫入 [{target_sheet.title}]" if target_sheet else "❌ 找不到分頁"
        })

        if not target_sheet:
            continue

        col_map = analyze_sheet_structure(target_sheet)
        target_row = find_target_row(target_sheet, data.date_day, col_map["date"])

        total_written += write_cell_if_valid(target_sheet, target_row, col_map.get("passenger"), to_clean_num(data.passenger_revenue))
        total_written += write_cell_if_valid(target_sheet, target_row, col_map.get("freight"), to_clean_num(data.freight_revenue))
        
        cc_formula = build_deduction_formula(data.credit_card_charge, data.credit_card_refund)
        total_written += write_cell_if_valid(target_sheet, target_row, col_map.get("credit"), cc_formula)

        bc_formula = build_deduction_formula(data.barcode_in, data.barcode_refund)
        total_written += write_cell_if_valid(target_sheet, target_row, col_map.get("barcode"), bc_formula)

        total_written += write_cell_if_valid(target_sheet, target_row, col_map.get("other"), to_clean_num(data.other_amount))
        total_written += write_cell_if_valid(target_sheet, target_row, col_map.get("remittance"), to_clean_num(data.remittance_total))

        success_count += 1

    status_box.update(label="🎉 辨識與 Excel 寫入完成！", state="complete")
    progress_bar.empty()

    out_stream = io.BytesIO()
    wb.save(out_stream)
    out_stream.seek(0)
    elapsed = time.time() - start_time

    st.balloons()
    st.success(f"✨ 處理完成！耗時 {elapsed:.1f} 秒，共成功處理 {success_count}/{total_tasks} 頁單據，填寫了 {total_written} 個儲存格！")

    # ----------------------------------------------------
    # 6. 會計平衡檢核儀表板
    # ----------------------------------------------------
    st.subheader("⚖️ 單據會計平衡勾稽核對表")
    if audit_records:
        df_audit = pd.DataFrame(audit_records)
        
        unbalanced_count = sum(1 for r in audit_records if "❌" in r["平衡狀態"])
        if unbalanced_count > 0:
            st.error(f"⚠️ 警告：共有 **{unbalanced_count}** 筆單據「自輸 - 總計」不等於 0，請依下方表格核對單據金額是否有誤！")
        else:
            st.success("🎯 太棒了！所有辨識成功的單據「自輸 - 總計」皆等於 0，會計平衡完全正確！")

        st.dataframe(
            df_audit,
            use_container_width=True,
            hide_index=True
        )

    st.download_button(
        label="📥 點擊下載已自動填寫完成的 Excel 報表 (.xlsx)",
        data=out_stream,
        file_name="台鐵解款單_彙總完成表.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True
    )
